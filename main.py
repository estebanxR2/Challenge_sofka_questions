import numpy as np
import pandas as pd
import openpyxl
import random
#path='C:\\Users\\jaire\\Downloads\\0_challenge_sofka\\DB.xlsx'
book= openpyxl.load_workbook('C:\\Users\\jaire\\Downloads\\0_challenge_sofka\\DB.xlsx', data_only=True) #poner la direccion dodne se encuentra la base de datos
hoja1=book['Hoja1'] #obtenemos la pestaña de la 1 pestaña tambien se puede con hoja=book.active
hoja2=book['Hoja2']
#print(book.sheetnames) #mirar el nombre de la pestañas del doc excel


puntaje=0
print('!!!!!!!!!!!!!!!!!! Challenge Preguntas y Respuestas !!!!!!!!!!!!!!!!!!!!!')
print()
instrucciones="""
-----------------------------------------------------INSTRUCCIONES--------------------------------------------------------
1) Se escogera una pregunta aleatoria la cual usted decidira una respuesta (A,B,C o D), si fallas perderas todo tu puntaje
pero si la respuesta es correcta continuaras con la siguiente ronda.
2) se le preguntara al usuario antes de seguir a una ronda si desea salir o continuar.
--------------------------------------------------------------------------------------------------------------------------
"""
print(instrucciones)
#---------------------------Datos usurario y puntaje-----------------------------------------------------------------------------------------------------------------------
name=str(input('Ingrese un nombre = '))
print()
print(name,"- puntaje ", puntaje)
print()


def categoria1(): #funcion para lanzar un numero aleatorio 
    numero = random.randint(2,6) #numero aleatorio de 2 a 6 de la categoria 1
    #print(numero)
    pregunta= hoja1.cell(row=numero, column=3).value # el numero aleatorio toma el numero de la fila y la columna queda predeterminado en 3
    print(pregunta)
    
    if numero is 2:
        respuestas=hoja2['B2:B5']
        for fila in respuestas:
            for respuesta in fila:
                 print(respuesta.value)
        
    elif numero is 3:
        respuestas=hoja2['B6:B9']
        for fila in respuestas:
            for respuesta in fila:
                 print(respuesta.value)
        
    elif numero is 4:
        respuestas=hoja2['B10:B13']
        for fila in respuestas:
            for respuesta in fila:
                 print(respuesta.value)
    
    elif numero is 5:
        respuestas=hoja2['B14:B17']
        for fila in respuestas:
            for respuesta in fila:
                 print(respuesta.value)
            
    elif numero is 6:
        respuestas=hoja2['B18:B21']
        for fila in respuestas:
            for respuesta in fila:
                 print(respuesta.value)
    print()        
    print('-Digite A,B,C o D la que crea correcta =')
    
def salir(): #funcion para salirse del programa 
    print('saliste del juego')
    exit()
    
    
d1=str(input('¿Quieres salir del juego? y/n= '))
print()
if d1=='y':
    salir()
if d1=='n':
    categoria1()

    
        

    
    
    
